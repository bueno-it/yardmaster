import csv
import io
import re
import json
from datetime import date, datetime

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.db.models import Q, Sum, Max

from .models import DailyPlan
from transport.models import Store, Load


def is_admin(user):
    return user.is_staff or user.is_superuser


@login_required
def plan_list(request):
    date_filter = request.GET.get('date', str(date.today()))
    search      = request.GET.get('search', '')

    plans = DailyPlan.objects.select_related('store').filter(date=date_filter).order_by('row_order', 'store_name_raw')

    if search:
        plans = plans.filter(
            Q(store__name__icontains=search) |
            Q(store_name_raw__icontains=search) |
            Q(store_id_raw__icontains=search) |
            Q(company__icontains=search)
        ).order_by('row_order', 'store_name_raw')

    total_ordered    = plans.aggregate(t=Sum('ordered'))['t'] or 0
    total_remaining  = plans.aggregate(t=Sum('remaining'))['t'] or 0
    total_picked     = sum(p.picked for p in plans)
    total_carryover  = total_remaining  # carryover = what's still remaining
    complete_count   = plans.filter(remaining=0).count()

    # Weekly total (Mon-Sun of selected date)
    from datetime import timedelta
    try:
        selected = datetime.strptime(date_filter, '%Y-%m-%d').date()
    except:
        selected = date.today()
    week_start = selected - timedelta(days=selected.weekday())
    week_end   = week_start + timedelta(days=6)
    week_ordered = DailyPlan.objects.filter(
        date__range=[week_start, week_end]
    ).aggregate(t=Sum('ordered'))['t'] or 0

    context = {
        'plans':           plans,
        'date_filter':     date_filter,
        'search':          search,
        'total_ordered':   total_ordered,
        'total_remaining': total_remaining,
        'total_carryover': total_carryover,
        'total_picked':    total_picked,
        'complete_count':  complete_count,
        'total_stores':    plans.count(),
        'week_ordered':    week_ordered,
        'week_start':      week_start,
        'week_end':        week_end,
        'is_admin':        is_admin(request.user),
    }
    return render(request, 'dailyplan/plan_list.html', context)


@login_required
def update_flag(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    data  = json.loads(request.body)
    plan  = get_object_or_404(DailyPlan, pk=data.get('id'))
    color = data.get('color', '')
    plan.flag_color = color
    plan.save(update_fields=['flag_color', 'updated_at'])
    return JsonResponse({'ok': True, 'color': color})


@login_required
def update_cell(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    data  = json.loads(request.body)
    plan  = get_object_or_404(DailyPlan, pk=data.get('id'))
    field = data.get('field')
    value = str(data.get('value', '')).strip()

    EDITABLE = ['hb_bd', 'doors', 'open_status', 'closed_status', 'company',
                'window', 'comments', 'mergers', 'loaders', 'check_field',
                'ordered', 'remaining', 'store_id_raw', 'store_name_raw']

    if field not in EDITABLE:
        return JsonResponse({'error': 'Not editable'}, status=400)

    if field in ('ordered', 'remaining'):
        try:
            value = int(value) if value else 0
        except ValueError:
            return JsonResponse({'error': 'Must be a number'}, status=400)

    setattr(plan, field, value)
    plan.save(update_fields=[field, 'updated_at'])
    _sync_to_yardmaster(plan)

    return JsonResponse({
        'ok': True,
        'picked': plan.picked,
        'progress_pct': plan.progress_pct,
        'is_complete': plan.is_complete,
        'is_rigid': 'rigid' in (plan.company or '').lower(),
    })


def _sync_to_yardmaster(plan):
    if not plan.store:
        return
    loads = Load.objects.filter(store=plan.store, date=plan.date)
    for load in loads:
        load.total_picked = plan.picked
        load.save(update_fields=['total_picked', 'updated_at'])


def _parse_date_val(val):
    if not val:
        return None
    if hasattr(val, 'date'):
        return val.date()
    if hasattr(val, 'year'):
        return val
    val = str(val).strip().split(' ')[0]
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            pass
    return None


def _to_int(v):
    try:
        return int(float(str(v).strip().replace(',', ''))) if v else 0
    except:
        return 0


# Map normalised header → model field
PLAN_HEADER_MAP = {
    'date':              'date',
    'hb/bd':             'hb_bd',
    'hb bd':             'hb_bd',
    'id':                'store_id_raw',
    'store':             'store_name_raw',
    'store name':        'store_name_raw',
    'store id':          'store_id_raw',
    'ordered':           'ordered',
    'remaining':         'remaining',
    'cases':             'ordered',
    'doors':             'doors',
    'open':              'open_status',
    'closed':            'closed_status',
    'company':           'company',
    'window':            'window',
    'comments':          'comments',
    'comment':           'comments',
    'mergers':           'mergers',
    'loaders':           'loaders',
    'pre load check':    'check_field',
    'pre-load check':    'check_field',
    'preload check':     'check_field',
    'check':             'check_field',
    'order':             'row_order',
    'pick carryover':    'pick_carryover',
    'carryover':         'pick_carryover',
    'amb pallets':       'pick_carryover',
    'wrap':              'mergers',
    'load':              'loaders',
}


def _read_file_rows(uploaded):
    """Read CSV or Excel and return (rows, error, is_new_format).

    is_new_format is True when the header row wasn't row 1 (a title/blank
    row sits above it) — that's our signal this came from the newer
    "Load plan for DD/MM/YYYY" export, which should only use a restricted
    set of columns even if it contains other populated fields.
    """
    fname = uploaded.name.lower()
    content = uploaded.read()

    if fname.endswith('.csv'):
        text = content.decode('utf-8-sig')
        rows = list(csv.DictReader(io.StringIO(text)))
        return rows, None, False

    elif fname.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

            # Find the (sheet, header_row) combo with the most matching headers.
            # We scan the first several rows of each sheet because some report
            # exports put a title (and a blank row) above the real header row.
            MAX_HEADER_SCAN_ROWS = 10
            best_ws         = None
            best_header_row = 1
            best_score      = -1
            for name in wb.sheetnames:
                ws = wb[name]
                max_row_to_scan = min(MAX_HEADER_SCAN_ROWS, ws.max_row or 1)
                for row_num in range(1, max_row_to_scan + 1):
                    try:
                        cell_row = next(ws.iter_rows(min_row=row_num, max_row=row_num))
                    except StopIteration:
                        continue
                    headers = [str(c.value).strip().lower() if c.value else '' for c in cell_row]
                    score = sum(1 for h in headers if h in PLAN_HEADER_MAP)
                    if score > best_score:
                        best_score      = score
                        best_ws         = ws
                        best_header_row = row_num

            if not best_ws or best_score <= 0:
                return None, 'No usable sheet found', False

            raw_headers = [str(c.value).strip() if c.value else f'col{i}'
                           for i, c in enumerate(next(best_ws.iter_rows(min_row=best_header_row, max_row=best_header_row)))]
            rows = []
            for excel_row in best_ws.iter_rows(min_row=best_header_row + 1, values_only=True):
                row = {}
                for h, v in zip(raw_headers, excel_row):
                    row[h] = str(v) if v is not None else ''
                rows.append(row)
            return rows, None, (best_header_row != 1)
        except ImportError:
            return None, 'openpyxl not installed', False
    else:
        return None, 'Only .csv, .xlsx supported', False


# Fields allowed for the newer "Load plan for DD/MM/YYYY" export format —
# everything else in that file (Comments, Check, Wrap, Load, Bay, etc.) is
# ignored even if populated.
NEW_FORMAT_ALLOWED_FIELDS = {
    'date', 'hb_bd', 'store_id_raw', 'store_name_raw', 'ordered', 'company', 'window',
}


@login_required
def import_plan(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    uploaded  = request.FILES.get('file')
    plan_date = request.POST.get('date', str(date.today()))

    if not uploaded:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    rows, err, is_new_format = _read_file_rows(uploaded)
    if err:
        return JsonResponse({'error': err}, status=400)
    if not rows:
        return JsonResponse({'error': 'Empty file'}, status=400)

    # Build column map from headers
    sample_headers = list(rows[0].keys())
    col_map = {}
    for h in sample_headers:
        key = h.strip().lower()
        if key in PLAN_HEADER_MAP:
            col_map[h] = PLAN_HEADER_MAP[key]

    created = updated = skipped = 0
    import_errors = []
    occurrence_counts = {}

    for row_idx, row in enumerate(rows):
        norm = {PLAN_HEADER_MAP[h.strip().lower()]: v.strip()
                for h, v in row.items()
                if h.strip().lower() in PLAN_HEADER_MAP and str(v).strip()}

        if is_new_format:
            norm = {k: v for k, v in norm.items() if k in NEW_FORMAT_ALLOWED_FIELDS}

        store_id = str(norm.get('store_id_raw', '')).strip().split('.')[0]  # remove .0 from floats
        if not store_id or store_id in ('None', 'nan', ''):
            skipped += 1
            continue

        # Determine date
        row_date = _parse_date_val(norm.get('date'))
        if not row_date:
            try:
                row_date = datetime.strptime(plan_date, '%Y-%m-%d').date()
            except:
                skipped += 1
                continue

        store_obj = Store.objects.filter(store_code=store_id).first()

        row_order_val = _to_int(norm.get('row_order', row_idx)) if norm.get('row_order') else row_idx
        hb_bd_val = norm.get('hb_bd', '')

        # Nth time this exact store+hb_bd combo appears in THIS file, for this date.
        # Using the occurrence count (instead of raw row position) as part of the
        # identity means reimporting the same file always matches the same
        # existing rows even if unrelated rows shifted position, while still
        # keeping genuine duplicate lines (same store+hb_bd twice) as separate rows.
        occ_key = (row_date, store_id, hb_bd_val)
        dup_seq_val = occurrence_counts.get(occ_key, 0)
        occurrence_counts[occ_key] = dup_seq_val + 1

        defaults = {
            'store':          store_obj,
            'store_name_raw': norm.get('store_name_raw', store_obj.name if store_obj else ''),
            'ordered':        _to_int(norm.get('ordered', 0)),
            'remaining':      _to_int(norm.get('remaining', norm.get('ordered', 0))),
            'hb_bd':          hb_bd_val,
            'doors':          norm.get('doors', ''),
            'open_status':    norm.get('open_status', ''),
            'closed_status':  norm.get('closed_status', ''),
            'company':        norm.get('company', ''),
            'window':         norm.get('window', ''),
            'comments':       norm.get('comments', ''),
            'mergers':        norm.get('mergers', ''),
            'loaders':        norm.get('loaders', ''),
            'check_field':    norm.get('check_field', ''),
            'row_order':      row_order_val,
            'dup_seq':        dup_seq_val,
        }

        try:
            obj, was_created = DailyPlan.objects.update_or_create(
                date=row_date, store_id_raw=store_id, hb_bd=hb_bd_val,
                dup_seq=dup_seq_val,
                defaults=defaults,
            )
            if was_created: created += 1
            else: updated += 1
        except Exception as e:
            skipped += 1
            import_errors.append({'row': row_idx + 2, 'store_id': store_id, 'error': str(e)})
            continue

    return JsonResponse({
        'ok': True,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': import_errors[:10],
    })


@login_required
def import_picking(request):
    """Parse Pick Remaining By Customer grouped CSV report."""
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    uploaded  = request.FILES.get('file')
    plan_date = request.POST.get('date', str(date.today()))

    if not uploaded:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    raw   = uploaded.read()
    fname = uploaded.name.lower()

    if fname.endswith(('.xlsx', '.xls')):
        return JsonResponse({'error': 'Picking report must be CSV. Please export as CSV from the picking system.'}, status=400)

    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        try:
            text = raw.decode('latin-1')
        except UnicodeDecodeError:
            text = raw.decode('utf-8', errors='replace')
    lines = text.splitlines()

    updated = not_found = 0
    current_store_id = None

    for line in lines:
        cols = [c.strip().strip('"') for c in line.split(',')]

        # Customer line: ",Customer:,,,318 - Monaghan 318, , ,"
        if len(cols) > 2 and cols[1].strip().lower() == 'customer:':
            raw = next((c for c in cols[2:] if c.strip()), '')
            match = re.match(r'(\d+)', raw.strip())
            current_store_id = match.group(1) if match else None
            continue

        # Totals line
        if cols[0].strip().lower() == 'totals' and current_store_id:
            non_empty = [c for c in cols if c.strip() not in ('', 'Totals')]
            if len(non_empty) >= 4:
                order_total  = _to_int(non_empty[1])
                picked_total = _to_int(non_empty[2])
                remaining    = _to_int(non_empty[3])
            elif len(non_empty) >= 3:
                order_total  = _to_int(non_empty[0])
                picked_total = _to_int(non_empty[1])
                remaining    = _to_int(non_empty[2])
            else:
                current_store_id = None
                continue

            try:
                matches = list(DailyPlan.objects.filter(date=plan_date, store_id_raw=current_store_id))
                if len(matches) == 1:
                    plan = matches[0]
                    plan.remaining = remaining
                    update_fields = ['remaining', 'updated_at']
                    if remaining == 0:
                        plan.closed_status = 'Closed'
                        update_fields.append('closed_status')
                    plan.save(update_fields=update_fields)
                    _sync_to_yardmaster(plan)
                    updated += 1
                elif len(matches) > 1:
                    not_found += 1
                else:
                    not_found += 1
            except DailyPlan.DoesNotExist:
                not_found += 1

            current_store_id = None

    return JsonResponse({'ok': True, 'updated': updated, 'not_found': not_found})


@login_required
def add_row(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    data     = json.loads(request.body)
    ref_id   = data.get('ref_id')
    position = data.get('position')

    if position not in ('above', 'below'):
        return JsonResponse({'error': 'Invalid position'}, status=400)

    ref_plan = get_object_or_404(DailyPlan, pk=ref_id)
    target_order = ref_plan.row_order if position == 'above' else ref_plan.row_order + 1

    # Shift every row at/after the target position up by 1 (highest first,
    # so we never momentarily collide with the unique constraint).
    to_shift = list(DailyPlan.objects.filter(
        date=ref_plan.date, row_order__gte=target_order
    ).order_by('-row_order'))
    for p in to_shift:
        p.row_order = p.row_order + 1
        p.save(update_fields=['row_order', 'updated_at'])

    new_plan_store_id = ''
    new_plan_hb_bd = ''
    max_dup = DailyPlan.objects.filter(
        date=ref_plan.date, store_id_raw=new_plan_store_id, hb_bd=new_plan_hb_bd
    ).aggregate(m=Max('dup_seq'))['m']
    new_dup_seq = (max_dup + 1) if max_dup is not None else 0

    new_plan = DailyPlan.objects.create(
        date=ref_plan.date,
        row_order=target_order,
        hb_bd=new_plan_hb_bd,
        store_id_raw=new_plan_store_id,
        store_name_raw='New store',
        ordered=0,
        remaining=0,
        dup_seq=new_dup_seq,
    )

    return JsonResponse({'ok': True, 'id': new_plan.pk})


@login_required
def delete_row(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    data = json.loads(request.body)
    plan = get_object_or_404(DailyPlan, pk=data.get('id'))
    plan.delete()
    return JsonResponse({'ok': True})


@login_required
def import_picking_v2(request):
    """Parse the newer 'Plan report' export: Customer Name, Customer No,
    Delivery Date, (blank), Order Total, Picked Total — one row per store
    per date, covering all stores (not just currently open ones).

    Rules (confirmed with the user):
      - Never touches Ordered, only Remaining.
      - Remaining = Order Total - Picked Total.
      - If a store has more than one row that date (split-load duplicates),
        only the row that already has Ordered > 0 (the "real" line) gets
        updated; the zeroed duplicate line is left untouched.
      - Closes the store (closed_status='Closed') when Remaining hits 0.
      - Does NOT auto-close stores just for being absent from the file.
    """
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    fname = uploaded.name.lower()
    if not fname.endswith('.csv'):
        return JsonResponse({'error': 'This report must be a CSV file.'}, status=400)

    raw = uploaded.read()
    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        try:
            text = raw.decode('latin-1')
        except UnicodeDecodeError:
            text = raw.decode('utf-8', errors='replace')

    all_rows = list(csv.reader(io.StringIO(text)))

    # Find the header row (the one that contains "Customer No") instead of
    # assuming a fixed position, since a title/blank row sits above it.
    header_idx = None
    for i, r in enumerate(all_rows):
        normalized = [c.strip().lower() for c in r]
        if 'customer no' in normalized:
            header_idx = i
            break

    if header_idx is None:
        return JsonResponse({'error': 'Could not find header row ("Customer No" column not found).'}, status=400)

    updated = not_found = skipped = 0

    for row in all_rows[header_idx + 1:]:
        if len(row) < 6:
            skipped += 1
            continue

        store_id = row[1].strip()
        if not store_id.isdigit():
            # Footer / summary rows (e.g. "Monday 20 July 2026 ... Page 1 of 1")
            # don't have a numeric Customer No, so they're skipped harmlessly.
            skipped += 1
            continue

        row_date = _parse_date_val(row[2])
        if not row_date:
            skipped += 1
            continue

        order_total  = _to_int(row[4])
        picked_total = _to_int(row[5])
        remaining    = max(0, order_total - picked_total)

        matches = list(DailyPlan.objects.filter(date=row_date, store_id_raw=store_id))

        target_plan = None
        if len(matches) == 1:
            target_plan = matches[0]
        elif len(matches) > 1:
            nonzero = [p for p in matches if p.ordered > 0]
            if len(nonzero) == 1:
                target_plan = nonzero[0]

        if not target_plan:
            not_found += 1
            continue

        target_plan.remaining = remaining
        update_fields = ['remaining', 'updated_at']
        if remaining == 0:
            target_plan.closed_status = 'Closed'
            update_fields.append('closed_status')
        target_plan.save(update_fields=update_fields)
        _sync_to_yardmaster(target_plan)
        updated += 1

    return JsonResponse({'ok': True, 'updated': updated, 'not_found': not_found, 'skipped': skipped})


@login_required
def export_plan(request):
    date_filter = request.GET.get('date', str(date.today()))
    plans = DailyPlan.objects.select_related('store').filter(date=date_filter)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="daily_plan_{date_filter}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Date', 'HB/BD', 'ID', 'Store', 'Ordered', 'Remaining', 'Picked',
                     'Pick Carryover', 'Progress %', 'Doors', 'Open', 'Closed',
                     'Company', 'Window', 'Comments', 'Mergers', 'Loaders', 'Check'])
    for p in plans:
        writer.writerow([
            p.date, p.hb_bd, p.store_id_raw, p.store_display,
            p.ordered, p.remaining, p.picked, p.pick_carryover,
            p.progress_pct,
            p.doors, p.open_status, p.closed_status, p.company, p.window,
            p.comments, p.mergers, p.loaders, p.check_field,
        ])
    return response
