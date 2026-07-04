import csv
import json
import io
from datetime import date, datetime

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_POST
from django.db.models import Q
from django.contrib import messages
from django.core.serializers.json import DjangoJSONEncoder

from .models import Load, Store, Driver, Company, CONSOLIDATED_CHOICES


def is_admin(user):
    return user.is_staff or user.is_superuser


# ── LOAD LIST ────────────────────────────────────────────────────────────────

@login_required
def load_list(request):
    loads = Load.objects.select_related('store', 'driver', 'company').all()

    date_filter   = request.GET.get('date', str(date.today()))
    day_filter    = request.GET.get('day', '')
    status_filter = request.GET.get('status', '')
    search        = request.GET.get('search', '')

    if date_filter:
        loads = loads.filter(date=date_filter)
    if day_filter:
        loads = loads.filter(day=day_filter)
    if status_filter:
        loads = loads.filter(consolidated=status_filter)
    if search:
        loads = loads.filter(
            Q(store__name__icontains=search) | Q(store_name__icontains=search) |
            Q(driver__name__icontains=search) | Q(driver_name__icontains=search) |
            Q(company__name__icontains=search) | Q(company_name__icontains=search) |
            Q(trailer_no__icontains=search) | Q(store_code__icontains=search)
        )

    total         = loads.count()
    loaded        = loads.filter(consolidated='Loaded').count()
    on_load       = loads.filter(consolidated='On the Load').count()
    total_pallets = sum(l.final_plt or 0 for l in loads)

    stores_qs    = Store.objects.filter(active=True)
    drivers_qs   = Driver.objects.filter(active=True)
    companies_qs = Company.objects.filter(active=True)

    context = {
        'loads':          loads,
        'stores':         stores_qs,
        'drivers':        drivers_qs,
        'companies':      companies_qs,
        'stores_json':    json.dumps(list(stores_qs.values('id', 'name', 'store_code')), cls=DjangoJSONEncoder),
        'drivers_json':   json.dumps(list(drivers_qs.values('id', 'name')), cls=DjangoJSONEncoder),
        'companies_json': json.dumps(list(companies_qs.values('id', 'name')), cls=DjangoJSONEncoder),
        'total':          total,
        'loaded':         loaded,
        'on_load':        on_load,
        'total_pallets':  total_pallets,
        'date_filter':    date_filter,
        'day_filter':     day_filter,
        'status_filter':  status_filter,
        'search':         search,
        'is_admin':       is_admin(request.user),
        'day_choices':    [d[0] for d in Load._meta.get_field('day').choices],
        'consolidated_choices': [c[0] for c in CONSOLIDATED_CHOICES if c[0]],
    }
    return render(request, 'transport/load_list.html', context)


# ── CELL UPDATE ──────────────────────────────────────────────────────────────

@login_required
def update_cell(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    data  = json.loads(request.body)
    load  = get_object_or_404(Load, pk=data.get('id'))
    field = data.get('field')
    value = str(data.get('value', '')).strip()

    EDITABLE_FIELDS = [
        'day', 'date', 'del_type', 'split_full', 'store_code',
        'final_plt', 'bay', 'time_arrive', 'departure_time', 'time_departed',
        'schedule_eta', 'date_departed', 'trailer_no',
        'consolidated', 'revised_eta', 'comment',
        'pre_checked_load', 'supervisor_checked',
    ]
    FK_FIELDS = {'store': Store, 'driver': Driver, 'company': Company}

    if field in FK_FIELDS:
        # value is the PK of the related object (or '' to clear)
        Model = FK_FIELDS[field]
        if value:
            obj = get_object_or_404(Model, pk=int(value))
            setattr(load, field, obj)
            # keep legacy text fields in sync
            if field == 'store':
                load.store_name = obj.name
                load.store_code = obj.store_code
            elif field == 'driver':
                load.driver_name = obj.name
            elif field == 'company':
                load.company_name = obj.name
            load.save()
            return JsonResponse({'ok': True, 'display': str(obj), 'code': getattr(obj, 'store_code', '')})
        else:
            setattr(load, field, None)
            load.save(update_fields=[field, 'updated_at'])
            return JsonResponse({'ok': True, 'display': '', 'code': ''})

    if field not in EDITABLE_FIELDS:
        return JsonResponse({'error': 'Field not editable'}, status=400)

    try:
        ft = Load._meta.get_field(field).get_internal_type()
        if ft == 'DateField':
            value = datetime.strptime(value, '%Y-%m-%d').date() if value else None
        elif ft == 'TimeField':
            value = datetime.strptime(value, '%H:%M').time() if value else None
        elif ft == 'IntegerField':
            value = int(value) if value else None
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

    setattr(load, field, value)
    load.save(update_fields=[field, 'updated_at'])
    return JsonResponse({'ok': True})


# ── FLAG UPDATE ───────────────────────────────────────────────────────────────

@login_required
def update_flag(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    data  = json.loads(request.body)
    load  = get_object_or_404(Load, pk=data.get('id'))
    color = data.get('color', '')
    valid = [c[0] for c in Load.ROW_COLOR_CHOICES]
    if color not in valid:
        return JsonResponse({'error': 'Invalid color'}, status=400)
    load.row_color = color
    load.save(update_fields=['row_color', 'updated_at'])
    return JsonResponse({'ok': True, 'color': color})


# ── ADD / DELETE ──────────────────────────────────────────────────────────────

@login_required
def add_load(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    data = json.loads(request.body)
    load = Load.objects.create(date=data.get('date') or date.today(), day=data.get('day', 'Mon'), status='Pending')
    return JsonResponse({'id': load.pk})


@login_required
def delete_load(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    data = json.loads(request.body)
    Load.objects.filter(pk__in=data.get('ids', [])).delete()
    return JsonResponse({'ok': True})


# ── EXPORT ────────────────────────────────────────────────────────────────────

@login_required
def export_csv(request):
    loads = Load.objects.select_related('store', 'driver', 'company').all()
    date_filter = request.GET.get('date', '')
    if date_filter:
        loads = loads.filter(date=date_filter)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="yardmaster_{date_filter or "all"}.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Day', 'Date', 'Del Seq', 'Split/Full', 'Store Name', 'Store ID',
        'Final Plt Count', 'Bay No', 'Time Arrived', 'Departure Time', 'Time Departed',
        'Driver Name', 'Company', 'Schedule ETA', 'Date Departed', 'Trailer No.',
        'Consolidated', 'Revised ETA', 'Comment', 'Pre-checked Load By', 'Supervisor Checked Load',
    ])
    for l in loads:
        writer.writerow([
            l.day, l.date, l.del_type, l.split_full,
            l.get_store_display(), l.get_store_code(),
            l.final_plt, l.bay, l.time_arrive, l.departure_time, l.time_departed,
            l.get_driver_display(), l.get_company_display(),
            l.schedule_eta, l.date_departed, l.trailer_no,
            l.consolidated, l.revised_eta, l.comment,
            l.pre_checked_load, l.supervisor_checked,
        ])
    return response


# ── IMPORT ────────────────────────────────────────────────────────────────────

HEADER_MAP = {
    'day': 'day', 'date': 'date',
    'del': 'del_type', 'del type': 'del_type', 'del seq': 'del_type', 'del  seq': 'del_type',
    'split': 'split_full', 'split/full': 'split_full', 'full/split': 'split_full',
    'store name': 'store_name', 'store': 'store_name',
    'store id': 'store_code', 'storeid': 'store_code', 'store no': 'store_code', 'id': 'store_code',
    'final plt': 'final_plt', 'final pallets': 'final_plt', 'pallets': 'final_plt',
    'planned plts': 'final_plt', 'planned pallets': 'final_plt',
    'bay': 'bay',
    'time arrive': 'time_arrive', 'arrival': 'time_arrive', 'time arrival': 'time_arrive', 'time arrived': 'time_arrive',
    'departure time': 'departure_time', 'departure': 'departure_time',
    'time departed': 'time_departed', 'time dep': 'time_departed',
    'driver name': 'driver_name', 'driver': 'driver_name',
    'drivers name': 'driver_name', "driver's name": 'driver_name',
    'company': 'company_name', 'haulier': 'company_name',
    'schedule eta': 'schedule_eta', 'sched eta': 'schedule_eta', 'eta': 'schedule_eta',
    'date departed': 'date_departed', 'date dep': 'date_departed',
    'date dep.': 'date_departed', 'departure date': 'date_departed',
    'trailer no': 'trailer_no', 'trailer no.': 'trailer_no', 'trailer': 'trailer_no',
    'consolidated': 'consolidated', 'revised eta': 'revised_eta',
    'wrapped and ready to load': 'consolidated',
    'wrapped and ready': 'consolidated',
    'wraped and ready to load': 'consolidated',
    'status': 'status', 'comment': 'comment', 'comments': 'comment', 'notes': 'comment',
    'pre-checked load': 'pre_checked_load', 'pre checked load': 'pre_checked_load',
    'pre-checked': 'pre_checked_load', 'prechecked': 'pre_checked_load',
    'supervisor checked load': 'supervisor_checked', 'supervisor checked': 'supervisor_checked',
    'supervisor': 'supervisor_checked', 'checked load': 'supervisor_checked',
}

# Sheet names to try in order when reading xlsx
PREFERRED_SHEETS = ['hourly', 'loads', 'transport', 'data', 'sheet1']


def _parse_date(val):
    if not val or str(val).strip() == '':
        return None
    val = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            pass
    return None


def _parse_time(val):
    if not val or str(val).strip() == '':
        return None
    val = str(val).strip()
    for fmt in ('%H:%M:%S', '%H:%M'):
        try:
            return datetime.strptime(val, fmt).time()
        except ValueError:
            pass
    return None


@login_required
def import_csv(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method == 'GET':
        return JsonResponse({'header_map': HEADER_MAP})

    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    filename = uploaded.name.lower()
    rows = []
    if filename.endswith('.csv'):
        text   = uploaded.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        rows   = list(reader)
    elif filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded, data_only=True)

            # Find the best sheet: prefer known names, else pick sheet with most columns matching HEADER_MAP
            ws = None
            sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

            # Try preferred sheet names first
            for preferred in PREFERRED_SHEETS:
                if preferred in sheet_names_lower:
                    ws = wb[sheet_names_lower[preferred]]
                    break

            # If not found, pick the sheet whose headers best match HEADER_MAP
            if ws is None:
                best_score = -1
                for sheet_name in wb.sheetnames:
                    candidate = wb[sheet_name]
                    try:
                        first_row = [str(c.value).strip().lower() if c.value else '' for c in next(candidate.iter_rows(min_row=1, max_row=1))]
                        score = sum(1 for h in first_row if h in HEADER_MAP)
                        if score > best_score:
                            best_score = score
                            ws = candidate
                    except StopIteration:
                        continue

            if ws is None:
                return JsonResponse({'error': 'No usable sheet found in the file'}, status=400)

            headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for excel_row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v) if v is not None else '' for v in excel_row])))
        except ImportError:
            return JsonResponse({'error': 'openpyxl not installed. Run: pip install openpyxl'}, status=400)
    else:
        return JsonResponse({'error': 'Only .csv, .xlsx, .xls supported'}, status=400)

    if not rows:
        return JsonResponse({'error': 'File is empty'}, status=400)

    sample_headers = list(rows[0].keys())
    col_map  = {h: HEADER_MAP[h.strip().lower()] for h in sample_headers if h.strip().lower() in HEADER_MAP}
    unmapped = [h for h in sample_headers if h.strip().lower() not in HEADER_MAP]

    if not col_map:
        return JsonResponse({'error': 'No recognisable columns found.', 'headers_found': sample_headers}, status=400)

    created = 0
    errors  = []
    for i, row in enumerate(rows, start=2):
        kwargs = {}
        for excel_h, model_f in col_map.items():
            raw = str(row.get(excel_h, '') or '').strip()
            if model_f in ('date', 'date_departed'):
                kwargs[model_f] = _parse_date(raw)
            elif model_f in ('time_arrive', 'departure_time', 'time_departed'):
                kwargs[model_f] = _parse_time(raw)
            elif model_f == 'final_plt':
                try:
                    kwargs[model_f] = int(float(raw)) if raw else None
                except ValueError:
                    kwargs[model_f] = None
            elif model_f == 'store_name':
                kwargs['store_name'] = raw
                # try to link FK
                if raw:
                    store_code = str(row.get('Store ID', row.get('store_code', row.get('Store No', ''))) or '').strip()
                    store, _ = Store.objects.get_or_create(name=raw, defaults={'store_code': store_code})
                    kwargs['store'] = store
            elif model_f == 'driver_name':
                kwargs['driver_name'] = raw
                if raw:
                    driver, _ = Driver.objects.get_or_create(name=raw)
                    kwargs['driver'] = driver
            elif model_f == 'company_name':
                kwargs['company_name'] = raw
                if raw:
                    company, _ = Company.objects.get_or_create(name=raw)
                    kwargs['company'] = company
            else:
                kwargs[model_f] = raw
        try:
            Load.objects.create(**kwargs)
            created += 1
        except Exception as e:
            errors.append({'row': i, 'error': str(e)})

    return JsonResponse({'ok': True, 'created': created, 'skipped': len(errors), 'unmapped_columns': unmapped, 'errors': errors[:10]})


# ── STORES CRUD ───────────────────────────────────────────────────────────────

@login_required
def store_list(request):
    if not is_admin(request.user):
        return redirect('load_list')
    stores = Store.objects.all()
    return render(request, 'transport/store_list.html', {'stores': stores, 'is_admin': True})

@login_required
def store_save(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    data = json.loads(request.body)
    pk   = data.get('id')
    name = data.get('name', '').strip()
    code = data.get('store_code', '').strip()
    active = data.get('active', True)
    if not name:
        return JsonResponse({'error': 'Name is required'}, status=400)
    if pk:
        obj = get_object_or_404(Store, pk=pk)
        obj.name = name; obj.store_code = code; obj.active = active; obj.save()
    else:
        obj = Store.objects.create(name=name, store_code=code, active=active)
    return JsonResponse({'ok': True, 'id': obj.pk, 'name': obj.name, 'store_code': obj.store_code, 'active': obj.active})

@login_required
def store_delete(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    data = json.loads(request.body)
    Store.objects.filter(pk__in=data.get('ids', [])).delete()
    return JsonResponse({'ok': True})


# ── DRIVERS CRUD ──────────────────────────────────────────────────────────────

@login_required
def driver_list(request):
    if not is_admin(request.user):
        return redirect('load_list')
    drivers = Driver.objects.all()
    return render(request, 'transport/driver_list.html', {'drivers': drivers, 'is_admin': True})

@login_required
def driver_save(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    data = json.loads(request.body)
    pk   = data.get('id')
    name = data.get('name', '').strip()
    active = data.get('active', True)
    if not name:
        return JsonResponse({'error': 'Name is required'}, status=400)
    if pk:
        obj = get_object_or_404(Driver, pk=pk)
        obj.name = name; obj.active = active; obj.save()
    else:
        obj = Driver.objects.create(name=name, active=active)
    return JsonResponse({'ok': True, 'id': obj.pk, 'name': obj.name, 'active': obj.active})

@login_required
def driver_delete(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    data = json.loads(request.body)
    Driver.objects.filter(pk__in=data.get('ids', [])).delete()
    return JsonResponse({'ok': True})


# ── COMPANIES CRUD ────────────────────────────────────────────────────────────

@login_required
def company_list(request):
    if not is_admin(request.user):
        return redirect('load_list')
    companies = Company.objects.all()
    return render(request, 'transport/company_list.html', {'companies': companies, 'is_admin': True})

@login_required
def company_save(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    data = json.loads(request.body)
    pk   = data.get('id')
    name = data.get('name', '').strip()
    active = data.get('active', True)
    if not name:
        return JsonResponse({'error': 'Name is required'}, status=400)
    if pk:
        obj = get_object_or_404(Company, pk=pk)
        obj.name = name; obj.active = active; obj.save()
    else:
        obj = Company.objects.create(name=name, active=active)
    return JsonResponse({'ok': True, 'id': obj.pk, 'name': obj.name, 'active': obj.active})

@login_required
def company_delete(request):
    if not is_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    data = json.loads(request.body)
    Company.objects.filter(pk__in=data.get('ids', [])).delete()
    return JsonResponse({'ok': True})


# ── API: list options for dropdowns ───────────────────────────────────────────

@login_required
def api_options(request):
    return JsonResponse({
        'stores':    list(Store.objects.filter(active=True).values('id', 'name', 'store_code')),
        'drivers':   list(Driver.objects.filter(active=True).values('id', 'name')),
        'companies': list(Company.objects.filter(active=True).values('id', 'name')),
    })
